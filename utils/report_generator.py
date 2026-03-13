import pandas as pd
import sqlite3
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime


def generate_report(session_id, section, subject, department, email):

    os.makedirs("reports", exist_ok=True)

    conn = sqlite3.connect("database/attendance.db")

    attendance = pd.read_sql_query(
        "SELECT roll, name, ip FROM attendance WHERE session_id=?",
        conn,
        params=(session_id,)
    )

    conn.close()

    # Create full class roll list
    roster = pd.DataFrame({
        "Roll Number": [str(i) for i in range(1, 151)]
    })

    attendance["roll"] = attendance["roll"].astype(str)

    roster["Name"] = roster["Roll Number"].map(
        attendance.set_index("roll")["name"]
    )

    roster["IP Address"] = roster["Roll Number"].map(
        attendance.set_index("roll")["ip"]
    )

    roster["Status"] = roster["Roll Number"].apply(
        lambda r: "Present" if r in attendance["roll"].values else "Absent"
    )

    roster["Section"] = section
    roster["Subject Code"] = subject
    roster["Department"] = department

    roster = roster[
        ["Roll Number","Name","Section","Subject Code","Department","IP Address","Status"]
    ]

    file_path = f"reports/attendance_{session_id}.xlsx"

    # Write Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        roster.to_excel(writer, startrow=8, index=False)

    wb = load_workbook(file_path)
    ws = wb.active

    # Add header information
    ws["A1"] = "Attendance Report"
    ws["A2"] = f"Professor: {email}"
    ws["A3"] = f"Subject Code: {subject}"
    ws["A4"] = f"Section: {section}"
    ws["A5"] = f"Department: {department}"
    ws["A6"] = f"Session ID: {session_id}"
    ws["A7"] = f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Highlight absentees
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in range(10, ws.max_row + 1):
        if ws[f"G{row}"].value == "Absent":
            ws[f"G{row}"].fill = red_fill

    wb.save(file_path)

    return file_path
